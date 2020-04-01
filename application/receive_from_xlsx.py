import numpy as np
from openpyxl.workbook import Workbook
import openpyxl
import xlrd
import numpy as np
import time
import os
import preprocessing.core_preprocess as prep

ORIGINAL_PATH = './data_gotten/'
DATA_FILE = 'data.xlsx'


SEQUENCE_LEN = 256

# 将存储数据的.xlsx文件转为(6, 256)手势序列numpy矩阵
def convert_to_numpy(data_file):
    book = xlrd.open_workbook(ORIGINAL_PATH+'{}'.format(data_file))
    # book = xlrd.open_workbook(data_file)

    table = book.sheet_by_index(0)  # 这里是引入第一个sheet，默认是引入第一个，要引入别的可以改那个数字

    # nrows=table.nrows
    # ncols=table.ncols

    start = 0
    end = SEQUENCE_LEN + start  # 这两个数是为了避开标题行，手动避的
    # rows=start-end

    list_values = []
    for i in range(6): # 列
        values = []
        try:
            for x in range(start, end): #行

                row = table.row_values(x)

                values.append(row[i])
        except IndexError:
            # 捕捉混乱表格引起的错误
            print('Error in original: ' + data_file)
        list_values.append(values)
    # print(list_values)
    data = np.array(list_values)
    return data

# 将存储了(6, ?)数据的表格整理为(6, 256)数据表格并(下方单元格上移)
# todo @author Victoire 连线均分算法
def intercept(sheet):
    num_row = sheet.max_row # 不算标题行
    proportion = 0.6  # 尾部删除比例或头部补充比例，因为尾部的无效数据比头部多，所以多删除一点

    # SEQUENCE_LEN = 256, 当数据的行数比256大的时候
    # 首尾删除数据，头部删除若干行，尾部删除若干行
    # 但是头部删除的行数和尾部删除的行数不一致，其服从 proportion=0.6的比例关系
    if num_row > SEQUENCE_LEN:
        # 分别计算首尾删除的行数
        devia = num_row - SEQUENCE_LEN
        down_remove = int(devia * proportion)
        up_remove = devia - down_remove

        # 掐头去尾
        if up_remove != 0:
            sheet.delete_rows(1, up_remove)
        if down_remove != 0:
            sheet.delete_rows(sheet.max_row - down_remove + 1, down_remove)

    # SEQUENCE_LEN = 256, 当数据的行数比256小的时候
    # 首尾补充数据，头部用第一行复制若干行，尾部用最后一行复制若干行
    # 但是尾部补充的行数和头部补充的行数不一致，其服从 proportion=0.6的比例关系
    elif num_row < SEQUENCE_LEN:
        # 分别计算首尾补充的行数
        devia = SEQUENCE_LEN - num_row
        up_insert = int(devia * proportion)
        down_insert = devia - up_insert

        # 通过复制第一行补充头部
        if up_insert != 0:
            sheet.insert_rows(1, up_insert)
            for i in range(1, up_insert + 1):
                for j in range(sheet.max_column):
                    sheet.cell(i, j+1).value = sheet.cell(2 + up_insert, j+1).value # 复制单元格数据

        # 通过复制最后一行补充尾部
        if down_insert != 0:
            sheet.insert_rows(sheet.max_row, down_insert)
            for i in range(sheet.max_row - down_insert, sheet.max_row):
                for j in range(sheet.max_column):
                 sheet.cell(i, j+1).value = sheet.cell(sheet.max_row, j+1).value # 复制单元格数据


def clear_invalid(sheet):
    sheet.delete_rows(sheet.max_row, 1)

# 返回 (6, 256) numpy 格式数据
def get_info():
    # time.sleep(3) # 等待 data_gotten 获取数据结束
    wb = openpyxl.load_workbook(ORIGINAL_PATH + DATA_FILE)
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    # clear_invalid(sheet)
    intercept(sheet)
    wb.save(ORIGINAL_PATH + DATA_FILE)
    time.sleep(0.7) # 等待保存
    data = convert_to_numpy(DATA_FILE)
    return data
