from openpyxl.workbook import Workbook
import openpyxl
import xlrd
import numpy as np
import os
import preprocessing.core_preprocess as prep

RAW_PATH = '../dataset/raw/'
ORIGINAL_PATH = '../dataset/original/'
NEW_PATH = '../dataset/new/'

SEQUENCE_LEN = 256

# 将存储数据的.xlsx文件转为(6, 256)手势序列numpy矩阵
def convert_to_numpy(data_file):
    book = xlrd.open_workbook(ORIGINAL_PATH+'{}'.format(data_file))
    # book = xlrd.open_workbook(data_file)

    table = book.sheet_by_index(0)  # 这里是引入第一个sheet，默认是引入第一个，要引入别的可以改那个数字

    # nrows=table.nrows
    # ncols=table.ncols

    start = 1
    end = SEQUENCE_LEN + start  # 这两个数是为了避开标题行，手动避的
    # rows=start-end

    list_values = []
    for i in range(1, 7): # 列
        values = []
        try:
            for x in range(start, end): #行

                row = table.row_values(x)

                values.append(row[i])
        except IndexError:
            # 捕捉混乱表格引起的错误
            print('Error in original: ' + data_file)
        list_values.append(values)
    # print(list_values)
    data = np.array(list_values)
    return data

# 将存储了(6, ?)数据的表格整理为(6, 256)数据表格并(下方单元格上移)
# todo @author Victoire 连线均分算法
def intercept(sheet):
    num_row = sheet.max_row - 1 # 不算标题行
    proportion = 0.6  # 尾部删除比例或头部补充比例，因为尾部的无效数据比头部多，所以多删除一点

    # SEQUENCE_LEN = 256, 当数据的行数比256大的时候
    # 首尾删除数据，头部删除若干行，尾部删除若干行
    # 但是头部删除的行数和尾部删除的行数不一致，其服从 proportion=0.6的比例关系
    if num_row > SEQUENCE_LEN:
        # 分别计算首尾删除的行数
        devia = num_row - SEQUENCE_LEN
        down_remove = int(devia * proportion)
        up_remove = devia - down_remove

        # 掐头去尾
        if up_remove != 0:
            sheet.delete_rows(2, up_remove)
        if down_remove != 0:
            sheet.delete_rows(sheet.max_row - down_remove + 1, down_remove)

    # SEQUENCE_LEN = 256, 当数据的行数比256小的时候
    # 首尾补充数据，头部用第一行复制若干行，尾部用最后一行复制若干行
    # 但是尾部补充的行数和头部补充的行数不一致，其服从 proportion=0.6的比例关系
    elif num_row < SEQUENCE_LEN:
        # 分别计算首尾补充的行数
        devia = SEQUENCE_LEN - num_row
        up_insert = int(devia * proportion)
        down_insert = devia - up_insert

        # 通过复制第一行补充头部
        if up_insert != 0:
            sheet.insert_rows(2, up_insert)
            for i in range(2, up_insert + 2):
                for j in range(sheet.max_column):
                    sheet.cell(i, j+1).value = sheet.cell(2 + up_insert, j+1).value # 复制单元格数据

        # 通过复制最后一行补充尾部
        if down_insert != 0:
            sheet.insert_rows(sheet.max_row, down_insert)
            for i in range(sheet.max_row - down_insert, sheet.max_row):
                for j in range(sheet.max_column):
                 sheet.cell(i, j+1).value = sheet.cell(sheet.max_row, j+1).value # 复制单元格数据


# 将原始数据集中的数据进行精简，变成convert_to_numpy能够接受的大小
def trim_xls(raw_path, original_path):
    dirs = os.listdir(raw_path)

    for dir in dirs:
        count_sequence = 0
        for file in os.listdir(raw_path + '{}'.format(dir)):
            try:
                wb = openpyxl.load_workbook(raw_path + dir + '/' + file)
            except:
                print(file) # 这里出现的问题是受损文件，即违规直接修改.txt后缀为.xlsx的文件
            sheet = wb.get_sheet_by_name(wb.sheetnames[0])
            sheet.delete_rows(1, 1) # intro
            sheet.delete_cols(1, 1) # adress
            sheet.delete_cols(5, 3) # wx, wy, wz
            sheet.delete_cols(8, 1) # T
            intercept(sheet)
            wb.save(original_path + dir + '/' + str(count_sequence) + '_' + dir + '.xlsx')

            # 下面代码发现运行途中，raw的表格中有几个存在空tuple，导致一系列问题
            # if count_sequence == 13:
            #     print('error in raw: ' + dir + '/' + file)
            count_sequence += 1
            wb.close()


# 将data写入 NEW_PATH/dir/dir_count_sequencce.xlsx中
def write_to_new(data, count_sequence, dir):
    outwb = Workbook()
    wo = outwb.active

    careerSheet = outwb.create_sheet('sheet1',0)   #创建的sheet

    for colnumber in range(1, 7):
        for rownumber in range(1, SEQUENCE_LEN + 1):
            careerSheet.cell(row=rownumber,column=colnumber).value =data[colnumber-1][rownumber-1]

    outwb.save(NEW_PATH + dir + '/' + dir + '_'  + str(count_sequence) + '.xlsx')

# 从原始xlsx数据集中读取数据，经过表格编辑、截取、预处理后，写到新数据集中
if __name__ == '__main__':

    trim_xls(RAW_PATH, ORIGINAL_PATH)

    datas = np.zeros(shape= [50, 6, 256], dtype= np.float32)
    dirs = os.listdir(ORIGINAL_PATH)

    try:
        for dir in dirs:
            count_sequence = 0
            for file in os.listdir(ORIGINAL_PATH+'{}'.format(dir)):
                datas[count_sequence] = convert_to_numpy(dir+'/'+file)
                datas[count_sequence] = prep.run_without_file(datas[count_sequence])
                write_to_new(datas[count_sequence], count_sequence, dir)
                count_sequence += 1
    except ValueError:

        # 遇到问题文件，先返回在 original 中的对应文件，然后根据文件名去 trim_xlsx 配置调试代码
        print('error in original: '+ dir + "/" + file)
