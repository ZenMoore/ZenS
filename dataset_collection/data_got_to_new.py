import  pandas as pd
import numpy as np
import os
import preprocessing.core_preprocess as prep
from openpyxl.workbook import Workbook
import openpyxl

SEQUENCE_LEN = 256
DIR_PATH = '../application/data_gotten'
OUT_PATH = '../dataset/due/'

# 将(6,?)的numpy转变为(6,256)的numpy
def intercept(data):
    length = data.shape[1]
    proportion = 0.6  # 尾部删除比例或头部补充比例，因为尾部的无效数据比头部多，所以多删除一点
    results = np.zeros(shape=[6, SEQUENCE_LEN])

    # SEQUENCE_LEN = 256, 当数据的行数比256大的时候
    # 首尾删除数据，头部删除若干行，尾部删除若干行
    # 但是头部删除的行数和尾部删除的行数不一致，其服从 proportion=0.6的比例关系
    if length > SEQUENCE_LEN:
        # 分别计算首尾删除的行数
        devia = length - SEQUENCE_LEN
        down_remove = int(devia * proportion)
        up_remove = devia - down_remove

        # 掐头去尾
        for i in range(0, SEQUENCE_LEN):
            for j in range(0, 6):
                results[j][i] = data[j][i+up_remove] # todo test

    # SEQUENCE_LEN = 256, 当数据的行数比256小的时候
    # 首尾补充数据，头部用第一行复制若干行，尾部用最后一行复制若干行
    # 但是尾部补充的行数和头部补充的行数不一致，其服从 proportion=0.6的比例关系
    elif length < SEQUENCE_LEN:
        # 分别计算首尾补充的行数
        devia = SEQUENCE_LEN - length
        up_insert = int(devia * proportion)
        down_insert = devia - up_insert

        # 通过复制第一行补充头部
        for i in range(0, up_insert):# todo test
            for j in range(0, 6):
                results[j][i] = data[j][0]

        for i in range(up_insert, length+up_insert):# todo test
            for j in range(0, 6):
                results[j][i] = data[j][i-up_insert]

        # 通过复制最后一行补充尾部
        for i in range(length+up_insert, SEQUENCE_LEN):# todo test
            for j in range(0, 6):
                results[j][i] = data[j][-1]
    else:
        for i in range(0, SEQUENCE_LEN):
            for j in range(0, 6):
                results[j][i] = data[j][i]

    return results


# 将(6, ?)的tsv文件读取为(6,?)的numpy
def convert_to_numpy(path):
    df = pd.DataFrame(pd.read_csv(path, sep='\t', header=None))
    length = df.shape[0] - 2  # 去掉最后两行数据，因为最后一行数据可能是无效数据(空)，倒数第二行可能不是一个tuple

    data = np.zeros(shape=[6, length])

    for i in range(0,length):
        for j in range(0, 6):
            data[j][i] = df.loc[i, j]

    return data

# 将(6, 256)的numpy存储到.xlsx中
# 将data写入 NEW_PATH/dir/dir_count_sequencce.xlsx中
def write_to_new(data, count_sequence, dir):
    outwb = Workbook()
    wo = outwb.active

    careerSheet = outwb.create_sheet('sheet1',0)   #创建的sheet

    for colnumber in range(1, 7):
        for rownumber in range(1, SEQUENCE_LEN + 1):
            careerSheet.cell(row=rownumber,column=colnumber).value =data[colnumber-1][rownumber-1]

    outwb.save(OUT_PATH + dir + '/' + dir + '_'  + str(count_sequence) + '.xlsx')

if __name__ == '__main__':
    type = 'croix'
    count = 0
    for file in os.listdir(DIR_PATH):
        count += 1
        data = convert_to_numpy(DIR_PATH+'/'+file)
        data = intercept(data)
        data = prep.run_without_file(data)
        write_to_new(data, count, type)